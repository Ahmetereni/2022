from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook('grades.xlsx')
ws=wb.active

filename=input("Enter filename: ")


with open (f"{filename}.txt", "r+") as myfile:
    data = myfile.read().splitlines()
    row_length=len(data)
    # print("rowlenth=",row_length)
    col_harf_sirasi=input("Enter column Coordinate(A,B,C,D...) : ")

    rowcoordinate= int(input("Enter Row Coordinate Number : "))-1
    index=0

    col_length=len(data[0].split())
    # print("len =",col_length)
    # myfile.write("stuff 12 32 543 231\n")
    col_sirasi=ord(col_harf_sirasi)-64

    for row in range(1+rowcoordinate,row_length+1+rowcoordinate):
        line_list= data[row-1-rowcoordinate].split()
        # print(line_list)
        for col in range(col_sirasi,col_length+col_sirasi):
            char=get_column_letter(col)
            # print(char,ws[char+str(row)].value)
            # char,ws[char+str(row)].value
            ws[char+str(row)].value=line_list[index]
            index+=1
            if index==col_length:
                index=0
    print("Done")
            
    wb.save("sonuc.xlsx")
     



# val=3
# rownumber=0
# collnumber=0
# for row in range(1,11):
#     for col in range(1,5):
#         char=get_column_letter(col)
#         print(ws[char+str(row)].value)
#     print()

# ws[f'{alfabe}{val}'].value=9
# wb.save('yeni.xlsx')
